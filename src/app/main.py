"""
Aplicación web Flask - Dashboard de detección de fraude en seguros.
Carga datos via web -> persiste en MySQL/SQLite -> analiza -> dashboard + Power BI.
"""
import os
import sys
import time
import threading
import traceback

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request, send_file
from flask_cors import CORS

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

load_dotenv()

from src.ingestion.load_data import load_all_from_directory, load_from_upload
from src.features.build_features import build_all_features, get_feature_columns
from src.rules.fraud_rules import apply_rules, get_rules_summary
from src.models.fraud_model import (
    train_supervised_model,
    train_anomaly_model,
    compute_hybrid_score,
    get_model_metrics_summary,
    predict_fraud_probability,
)
from src.nlp.text_analysis import get_similarity_scores_by_id, generate_text_summary
from src.app.dashboard_service import (
    apply_dashboard_filters,
    build_dashboard_payload,
    get_filter_options,
    params_from_request,
)
from src.explainability.explain_score import explain_single_case, generate_executive_summary
from src.utils.dataframe_columns import ensure_str_columns, normalize_datasets_columns
from src.ai_agent.claims_agent import ClaimsAgent
from src.ai_agent.openai_client import is_openai_configured, get_openai_model
from src.app.powerbi_export import export_to_powerbi, export_csv_for_powerbi
from src.app.vercel_bootstrap import bootstrap_vercel_demo, is_vercel_runtime
from src.pipeline.run_full_analysis import execute_full_pipeline
from src.db.config import get_engine, test_connection
from src.db.repository import (
    init_database,
    save_all_datasets,
    load_all_datasets,
    update_siniestros_scores,
    save_analysis_run,
    get_analysis_history,
    get_db_stats,
    save_dataframe,
)

app = Flask(__name__, template_folder="templates", static_folder="static")
CORS(app)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024

UPLOAD_FOLDER = os.path.join("data", "raw")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app_state = {
    "datasets": {},
    "df_features": None,
    "df_scored": None,
    "model_results": None,
    "anomaly_results": None,
    "nlp_results": None,
    "agent": None,
    "dashboard_snapshot": None,
    "model_snapshot": None,
    "dashboard_last_payload": None,
    "pipeline_status": "idle",
}

_db_save_lock = threading.Lock()


@app.before_request
def _ensure_vercel_demo_loaded():
    """Carga bundle de analisis en el primer request (Vercel)."""
    if is_vercel_runtime() and app_state.get("df_scored") is None:
        try:
            bootstrap_vercel_demo(app_state)
        except Exception as exc:
            print(f"[Vercel] Error cargando bundle: {exc}")


def _reset_pipeline_state():
    """Invalida resultados de análisis previos al cargar datos nuevos."""
    app_state["df_features"] = None
    app_state["df_scored"] = None
    app_state["model_results"] = None
    app_state["anomaly_results"] = None
    app_state["nlp_results"] = None
    app_state["agent"] = None
    app_state["dashboard_snapshot"] = None
    app_state["model_snapshot"] = None
    app_state["dashboard_last_payload"] = None
    app_state["pipeline_status"] = "idle"


def _build_agent_context() -> dict:
    """Contexto persistente para respuestas del agente (dashboard + modelo)."""
    return {
        "dashboard_snapshot": app_state.get("dashboard_snapshot"),
        "model_snapshot": app_state.get("model_snapshot"),
        "dashboard_last_payload": app_state.get("dashboard_last_payload"),
    }


def _merge_uploaded_tables(new_tables: dict) -> None:
    """
    Incorpora tablas del archivo subido.
    Workbook completo (3+ tablas conocidas con siniestros) → reemplaza el dataset.
    Hoja/CSV suelto → actualiza solo esas tablas (conserva el resto en memoria).
    """
    if not new_tables:
        return

    known = {"siniestros", "polizas", "asegurados", "vehiculos", "proveedores", "documentos"}
    known_in_file = set(new_tables.keys()) & known
    is_full_workbook = "siniestros" in new_tables and len(known_in_file) >= 3

    if is_full_workbook:
        app_state["datasets"] = new_tables
        return

    if not app_state.get("datasets"):
        app_state["datasets"] = {}

    for name, df in new_tables.items():
        app_state["datasets"][name] = df


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/health")
def health():
    """Comprobacion rapida de despliegue (sin pipeline)."""
    return jsonify({
        "status": "ok",
        "vercel": is_vercel_runtime(),
        "pipeline_ready": app_state.get("df_scored") is not None,
    })


# ── Database endpoints ───────────────────────────────────────────────

@app.route("/api/deployment-info")
def deployment_info():
    """Metadatos del entorno (local vs Vercel)."""
    from src.pipeline.run_full_analysis import load_vercel_bundle

    bundle = load_vercel_bundle() if is_vercel_runtime() else {}
    manifest = bundle.get("manifest", {})
    return jsonify({
        "vercel": is_vercel_runtime(),
        "pipeline_ready": app_state.get("df_scored") is not None,
        "openai_configured": is_openai_configured(),
        "openai_model": get_openai_model() if is_openai_configured() else None,
        "analysis_flow": (
            "build: datos → pipeline → bundle JSON + CSV; "
            "runtime: dashboard + agente OpenAI sobre ese análisis"
        ) if is_vercel_runtime() else "local: carga/sintéticos → pipeline en /api/run-pipeline",
        "manifest": manifest,
        "records": manifest.get("total_records") or (
            len(app_state["df_scored"]) if app_state.get("df_scored") is not None else 0
        ),
    })


@app.route("/api/db-status")
def db_status():
    if is_vercel_runtime():
        return jsonify({
            "status": "ok",
            "type": "In-memory (Vercel)",
            "host": "demo",
            "message": "En Vercel los datos vienen del CSV embebido; no hay SQLite persistente.",
        })
    conn = test_connection()
    if conn["status"] == "ok":
        stats = get_db_stats()
        conn["stats"] = stats
        conn["history"] = get_analysis_history()
    return jsonify(conn)


@app.route("/api/db-init", methods=["POST"])
def db_init():
    try:
        init_database()
        return jsonify({"status": "ok", "message": "Tablas creadas en la base de datos"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Template download ─────────────────────────────────────────────────

@app.route("/api/download-template")
def download_template():
    """Generate and return an Excel template with headers and descriptions for each entity."""
    import io

    templates = {
        "siniestros": [
            "id_siniestro", "id_poliza", "id_asegurado", "ramo", "cobertura",
            "fecha_ocurrencia", "fecha_reporte", "monto_reclamado", "monto_estimado",
            "monto_pagado", "estado", "sucursal", "descripcion", "documentos_completos",
            "beneficiario", "dias_desde_inicio_poliza", "dias_desde_fin_poliza",
            "dias_entre_ocurrencia_reporte", "historial_siniestros_asegurado",
            "etiqueta_fraude_simulada",
        ],
        "polizas": [
            "id_poliza", "id_asegurado", "ramo", "fecha_inicio", "fecha_fin",
            "prima", "suma_asegurada", "deducible", "canal_venta", "ciudad", "estado_poliza",
        ],
        "asegurados": [
            "id_asegurado", "segmento", "antiguedad_anos", "ciudad",
            "numero_polizas", "reclamos_ultimos_12m", "mora_actual", "score_cliente",
        ],
        "proveedores": [
            "id_proveedor", "tipo", "ciudad", "reclamos_asociados",
            "monto_promedio_reclamado", "casos_observados", "antiguedad_anos",
        ],
        "documentos": [
            "id_documento", "id_siniestro", "tipo_documento", "entregado",
            "legible", "fecha_emision", "inconsistencia_detectada", "observacion",
        ],
    }

    guia = pd.DataFrame([
        {"Tabla": "siniestros", "Campos_clave": "id_siniestro, id_poliza, id_asegurado", "Propósito": "Tabla principal del análisis antifraude"},
        {"Tabla": "polizas", "Campos_clave": "id_poliza, id_asegurado", "Propósito": "Vigencia y suma asegurada"},
        {"Tabla": "asegurados", "Campos_clave": "id_asegurado", "Propósito": "Historial y perfil del asegurado"},
        {"Tabla": "proveedores", "Campos_clave": "id_proveedor", "Propósito": "Riesgo por proveedor/beneficiario"},
        {"Tabla": "documentos", "Campos_clave": "id_documento, id_siniestro", "Propósito": "Consistencia documental"},
    ])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet, cols in templates.items():
            pd.DataFrame(columns=cols).to_excel(writer, sheet_name=sheet, index=False)
        guia.to_excel(writer, sheet_name="Guia", index=False)

        wb = writer.book
        from openpyxl.styles import Alignment, Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 48)

    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        as_attachment=True,
        download_name="plantilla_dataset_fraudia.xlsx",
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Siniestros
        pd.DataFrame(columns=[
            "id_siniestro", "id_poliza", "id_asegurado", "id_vehiculo", "id_conductor",
            "ramo", "cobertura", "fecha_ocurrencia", "fecha_reporte",
            "monto_reclamado", "monto_estimado", "monto_pagado", "estado", "sucursal",
            "descripcion", "documentos_completos", "id_proveedor", "beneficiario",
            "dias_desde_inicio_poliza", "dias_desde_fin_poliza",
            "dias_entre_ocurrencia_reporte", "historial_siniestros_asegurado",
            "etiqueta_fraude_simulada",
        ]).to_excel(writer, sheet_name="siniestros", index=False)

        # Polizas
        pd.DataFrame(columns=[
            "id_poliza", "id_asegurado", "ramo", "fecha_inicio", "fecha_fin",
            "prima", "suma_asegurada", "deducible", "canal_venta", "ciudad",
            "estado_poliza",
        ]).to_excel(writer, sheet_name="polizas", index=False)

        # Asegurados
        pd.DataFrame(columns=[
            "id_asegurado", "segmento", "antiguedad_anos", "ciudad",
            "numero_polizas", "reclamos_ultimos_12m", "mora_actual", "score_cliente",
        ]).to_excel(writer, sheet_name="asegurados", index=False)

        # Vehiculos
        pd.DataFrame(columns=[
            "id_vehiculo", "id_asegurado", "placa", "chasis", "motor",
            "marca", "modelo", "ano", "color", "tipo",
        ]).to_excel(writer, sheet_name="vehiculos", index=False)

        # Proveedores
        pd.DataFrame(columns=[
            "id_proveedor", "nombre_proveedor", "tipo", "ciudad",
            "reclamos_asociados", "monto_promedio_reclamado", "casos_observados",
            "en_lista_restrictiva", "antiguedad_anos",
        ]).to_excel(writer, sheet_name="proveedores", index=False)

        # Documentos
        pd.DataFrame(columns=[
            "id_documento", "id_siniestro", "tipo_documento", "entregado",
            "legible", "fecha_emision", "inconsistencia_detectada", "observacion",
        ]).to_excel(writer, sheet_name="documentos", index=False)

        # --- Instrucciones sheet ---
        instrucciones = pd.DataFrame({
            "Hoja": [
                "siniestros", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                "polizas", "", "", "", "", "", "", "", "", "", "",
                "asegurados", "", "", "", "", "", "", "",
                "vehiculos", "", "", "", "", "", "", "", "", "",
                "proveedores", "", "", "", "", "", "", "", "",
                "documentos", "", "", "", "", "", "", "",
            ],
            "Campo": [
                "id_siniestro", "id_poliza", "id_asegurado", "id_vehiculo", "id_conductor",
                "ramo", "cobertura", "fecha_ocurrencia", "fecha_reporte",
                "monto_reclamado", "monto_estimado", "monto_pagado", "estado", "sucursal",
                "descripcion", "documentos_completos", "id_proveedor", "beneficiario",
                "dias_desde_inicio_poliza", "dias_desde_fin_poliza",
                "dias_entre_ocurrencia_reporte", "historial_siniestros_asegurado",
                "etiqueta_fraude_simulada",
                "id_poliza", "id_asegurado", "ramo", "fecha_inicio", "fecha_fin",
                "prima", "suma_asegurada", "deducible", "canal_venta", "ciudad",
                "estado_poliza",
                "id_asegurado", "segmento", "antiguedad_anos", "ciudad",
                "numero_polizas", "reclamos_ultimos_12m", "mora_actual", "score_cliente",
                "id_vehiculo", "id_asegurado", "placa", "chasis", "motor",
                "marca", "modelo", "ano", "color", "tipo",
                "id_proveedor", "nombre_proveedor", "tipo", "ciudad",
                "reclamos_asociados", "monto_promedio_reclamado", "casos_observados",
                "en_lista_restrictiva", "antiguedad_anos",
                "id_documento", "id_siniestro", "tipo_documento", "entregado",
                "legible", "fecha_emision", "inconsistencia_detectada", "observacion",
            ],
            "Descripcion": [
                "Identificador unico del siniestro (ej: SIN-000001)",
                "ID de la poliza asociada (ej: POL-000001)",
                "ID anonimo del asegurado (ej: ASE-000001)",
                "ID del vehiculo, si aplica (ej: VEH-000001)",
                "ID del conductor al momento del siniestro",
                "Vehiculos, Salud, Vida, Generales, Hogar",
                "Choque, Robo, Atencion Medica, Incendio, Dano, etc.",
                "Fecha del evento (YYYY-MM-DD)",
                "Fecha de notificacion (YYYY-MM-DD)",
                "Valor solicitado por el asegurado o proveedor",
                "Valor estimado por la aseguradora",
                "Valor pagado, si aplica",
                "Reserva, Pago Total, Pago Parcial, Anticipo, Negativa, Cierre Sin Consecuencia, Liquidado",
                "Sucursal del siniestro",
                "Texto libre describiendo el reclamo",
                "Si/No - Indicador si la documentacion esta completa",
                "ID del proveedor/beneficiario (ej: PROV-000001)",
                "Taller, clinica, perito u otro beneficiario",
                "Dias entre inicio de poliza y siniestro",
                "Dias entre fin de poliza y siniestro",
                "Diferencia en dias entre ocurrencia y reporte",
                "Numero de siniestros previos del asegurado",
                "0 o 1 - Solo para entrenamiento/evaluacion del modelo",
                "ID unico de la poliza (ej: POL-000001)",
                "ID del asegurado titular",
                "Vehiculos, Salud, Vida, Generales, Hogar",
                "Fecha inicio vigencia (YYYY-MM-DD)",
                "Fecha fin vigencia (YYYY-MM-DD)",
                "Prima de la poliza (valor numerico)",
                "Suma asegurada total",
                "Deducible aplicable",
                "Agente, Broker, Digital, Directo",
                "Ciudad de la poliza",
                "Activa, Cancelada, Vencida",
                "ID anonimo del asegurado (ej: ASE-000001)",
                "Premium, Estandar, Basico",
                "Antiguedad del cliente en anos",
                "Ciudad del asegurado",
                "Cantidad de polizas activas",
                "Reclamos en los ultimos 12 meses",
                "1 si tiene mora, 0 si no",
                "Score de cliente simulado (0-100)",
                "ID unico del vehiculo (ej: VEH-000001)",
                "ID del asegurado propietario",
                "Numero de placa del vehiculo",
                "Numero de chasis / VIN",
                "Numero de motor",
                "Marca del vehiculo (Toyota, Hyundai, etc.)",
                "Modelo del vehiculo (Corolla, Tucson, etc.)",
                "Ano de fabricacion (ej: 2020)",
                "Color del vehiculo",
                "Sedan, SUV, Pickup, Camioneta, etc.",
                "ID unico del proveedor (ej: PROV-000001)",
                "Nombre del proveedor o taller",
                "Taller, Clinica, Perito, Abogado",
                "Ciudad del proveedor",
                "Cantidad de reclamos asociados",
                "Monto promedio reclamado por este proveedor",
                "Numero de casos observados/flaggeados",
                "1 si esta en lista restrictiva, 0 si no",
                "Antiguedad del proveedor en anos",
                "ID unico del documento (ej: DOC-000001)",
                "ID del siniestro al que pertenece",
                "Denuncia, Factura, Informe Medico, Peritaje, etc.",
                "Si/No - Si fue entregado",
                "Si/No - Si es legible",
                "Fecha de emision del documento (YYYY-MM-DD)",
                "Texto describiendo inconsistencia detectada, si aplica",
                "Observaciones adicionales",
            ],
            "Tipo": [
                "Texto (clave)", "Texto (FK)", "Texto (FK)", "Texto (FK)", "Texto",
                "Texto", "Texto", "Fecha", "Fecha",
                "Numero", "Numero", "Numero", "Texto", "Texto",
                "Texto largo", "Si/No", "Texto (FK)", "Texto",
                "Numero entero", "Numero entero",
                "Numero entero", "Numero entero",
                "0 o 1",
                "Texto (clave)", "Texto (FK)", "Texto", "Fecha", "Fecha",
                "Numero", "Numero", "Numero", "Texto", "Texto",
                "Texto",
                "Texto (clave)", "Texto", "Numero entero", "Texto",
                "Numero entero", "Numero entero", "0 o 1", "Numero decimal",
                "Texto (clave)", "Texto (FK)", "Texto", "Texto", "Texto",
                "Texto", "Texto", "Numero entero", "Texto", "Texto",
                "Texto (clave)", "Texto", "Texto", "Texto",
                "Numero entero", "Numero decimal", "Numero entero",
                "0 o 1", "Numero entero",
                "Texto (clave)", "Texto (FK)", "Texto", "Si/No",
                "Si/No", "Fecha", "Texto largo", "Texto largo",
            ],
            "Requerido": [
                "Si", "Si", "Si", "No", "No",
                "Si", "Si", "Si", "Si",
                "Si", "No", "No", "Si", "No",
                "Si", "Si", "No", "No",
                "No", "No",
                "No", "No",
                "No",
                "Si", "Si", "Si", "Si", "Si",
                "No", "No", "No", "No", "No",
                "No",
                "Si", "No", "No", "No",
                "No", "No", "No", "No",
                "Si", "Si", "No", "No", "No",
                "No", "No", "No", "No", "No",
                "Si", "No", "No", "No",
                "No", "No", "No",
                "No", "No",
                "Si", "Si", "No", "No",
                "No", "No", "No", "No",
            ],
        })
        instrucciones.to_excel(writer, sheet_name="Instrucciones", index=False)

        wb = writer.book
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        ws_inst = wb["Instrucciones"]
        section_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        section_font = Font(bold=True, color="1E40AF")
        for row in ws_inst.iter_rows(min_row=2, max_col=1):
            if row[0].value:
                for cell in ws_inst[row[0].row]:
                    cell.fill = section_fill
                    if cell.column == 1:
                        cell.font = section_font

    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        as_attachment=True,
        download_name="plantilla_dataset_fraudia.xlsx",
    )


# ── Upload & load endpoints ──────────────────────────────────────────

@app.route("/api/upload", methods=["POST"])
def upload_dataset():
    try:
        if "file" not in request.files:
            return jsonify({"error": "No se envio archivo"}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "Nombre de archivo vacio"}), 400

        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in (".csv", ".xlsx", ".xls"):
            return jsonify({"error": "Formato no soportado. Use CSV o Excel (.xlsx)."}), 400

        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        from src.ingestion.load_data import load_file_to_tables, validate_datasets

        new_tables = {
            name: ensure_str_columns(df)
            for name, df in load_file_to_tables(filepath, file.filename).items()
        }
        if not new_tables:
            return jsonify({"error": "El archivo no contiene datos válidos (hojas vacías)."}), 400

        _merge_uploaded_tables(new_tables)
        _reset_pipeline_state()

        validation = validate_datasets(app_state["datasets"])
        tables_info = {
            name: {"rows": len(df), "columns": len(df.columns), "cols": list(df.columns)}
            for name, df in app_state["datasets"].items()
        }

        def _save_to_db(datasets):
            if "siniestros" not in datasets:
                return
            with _db_save_lock:
                try:
                    init_database()
                    save_all_datasets({k: v.copy() for k, v in datasets.items()})
                except Exception:
                    pass

        if validation["has_siniestros"]:
            datasets_copy = {k: v.copy() for k, v in app_state["datasets"].items()}
            threading.Thread(target=_save_to_db, args=(datasets_copy,), daemon=True).start()
            db_info = test_connection()
            db_msg = f" (guardando en {db_info.get('type', 'BD')} en segundo plano)"
        else:
            db_msg = ""

        msg = f"'{file.filename}' cargado ({', '.join(new_tables.keys())})."
        if len(app_state["datasets"]) > len(new_tables):
            msg += f" Dataset en memoria: {', '.join(app_state['datasets'].keys())}."
        msg += db_msg

        status = "success" if validation["has_siniestros"] else "warning"
        return jsonify({
            "status": status,
            "message": msg,
            "tables": tables_info,
            "has_siniestros": validation["has_siniestros"],
            "warnings": validation["warnings"],
            "loaded_tables": list(new_tables.keys()),
        })
    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/load-synthetic", methods=["POST"])
def load_synthetic():
    try:
        from src.ingestion.generate_synthetic import main as generate_data

        synthetic_dir = os.path.join("data", "synthetic")
        seed_used = generate_data()

        app_state["datasets"] = load_all_from_directory(synthetic_dir)
        _reset_pipeline_state()

        tables_info = {
            name: {"rows": len(df), "columns": len(df.columns)}
            for name, df in app_state["datasets"].items()
        }

        def _save_synthetic_db(datasets):
            try:
                init_database()
                save_all_datasets(datasets)
            except Exception:
                pass

        datasets_copy = {k: v.copy() for k, v in app_state["datasets"].items()}
        threading.Thread(target=_save_synthetic_db, args=(datasets_copy,), daemon=True).start()

        db_info = test_connection()
        db_msg = f" (guardando en {db_info.get('type', 'BD')} en segundo plano)"

        return jsonify({
            "status": "success",
            "message": f"Datos sinteticos generados (semilla {seed_used}){db_msg}",
            "seed": seed_used,
            "tables": tables_info,
        })
    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/load-from-db", methods=["POST"])
def load_from_db():
    """Load datasets already stored in the database (for when data was uploaded previously)."""
    try:
        datasets = normalize_datasets_columns(load_all_datasets())
        if not datasets or "siniestros" not in datasets:
            return jsonify({"error": "No hay datos en la base de datos. Suba un archivo primero."}), 400

        app_state["datasets"] = datasets
        _reset_pipeline_state()

        tables_info = {
            name: {"rows": len(df), "columns": len(df.columns)}
            for name, df in datasets.items()
        }
        db_info = test_connection()

        return jsonify({
            "status": "success",
            "message": f"Datos cargados desde {db_info.get('type', 'DB')} ({db_info.get('host', 'local')})",
            "tables": tables_info,
            "db_type": db_info.get("type", "unknown"),
            "has_siniestros": True,
        })
    except Exception as e:
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── Analysis pipeline ────────────────────────────────────────────────

@app.route("/api/run-pipeline", methods=["POST"])
def run_pipeline():
    try:
        if is_vercel_runtime():
            boot = bootstrap_vercel_demo(app_state)
            return jsonify({
                "status": "success",
                "message": (
                    "En Vercel el análisis completo se ejecuta al desplegar (build): "
                    "carga/generación de datos → reglas → ML → dashboard. "
                    "El chatbot OpenAI explica esos resultados. Para recalcular: redeploy."
                ),
                "total_records": len(app_state["df_scored"]) if app_state.get("df_scored") is not None else 0,
                "vercel": True,
                "bootstrap": boot,
            })

        if not app_state["datasets"] or "siniestros" not in app_state["datasets"]:
            return jsonify({"error": "No hay datos cargados. Suba un archivo o cargue datos sinteticos."}), 400

        t_start = time.time()
        app_state["pipeline_status"] = "running"

        result = execute_full_pipeline(app_state["datasets"])
        df_scored = result["df_scored"]
        app_state["datasets"] = result["datasets"]
        app_state["df_features"] = result["df_features"]
        app_state["df_scored"] = df_scored
        app_state["model_results"] = result["model_results"]
        app_state["anomaly_results"] = result["anomaly_results"]
        app_state["nlp_results"] = result["nlp_results"]
        app_state["dashboard_snapshot"] = result["dashboard_snapshot"]
        app_state["dashboard_last_payload"] = result["dashboard_last_payload"]
        app_state["model_snapshot"] = result["model_snapshot"]

        sem_counts = df_scored["semaforo_final"].value_counts()
        db_payload = {
            "total": len(df_scored),
            "rojos": int(sem_counts.get("Rojo", 0)),
            "amarillos": int(sem_counts.get("Amarillo", 0)),
            "verdes": int(sem_counts.get("Verde", 0)),
            "score_prom": float(df_scored["score_hibrido"].mean()),
            "auc": result.get("auc_roc"),
            "anomalias": result["anomaly_results"]["n_anomalies"] if result.get("anomaly_results") else None,
            "duracion": time.time() - t_start,
        }

        def _persist_scores(df, payload):
            try:
                update_siniestros_scores(df)
                save_analysis_run(**payload)
            except Exception:
                pass

        threading.Thread(target=_persist_scores, args=(df_scored.copy(), db_payload), daemon=True).start()

        app_state["agent"] = ClaimsAgent(df_scored, extra_context=_build_agent_context())
        app_state["pipeline_status"] = "completed"

        results = {
            "status": "success",
            "steps": result["steps"] + [{"step": "Guardado en BD (en segundo plano)", "status": "ok"}],
            "executive_summary": result["executive_summary"],
            "total_records": result["total_records"],
            "duration_seconds": result["duration_seconds"],
        }
        return jsonify(results)
    except Exception as e:
        app_state["pipeline_status"] = "error"
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── Dashboard & case endpoints ───────────────────────────────────────

@app.route("/api/dashboard-filters")
def dashboard_filters():
    df = app_state.get("df_scored")
    if df is None:
        return jsonify({"error": "Pipeline no ejecutado"}), 400
    return jsonify(get_filter_options(df))


@app.route("/api/dashboard-data")
def dashboard_data():
    df = app_state.get("df_scored")
    if df is None:
        return jsonify({"error": "Pipeline no ejecutado"}), 400

    params = params_from_request(request)
    df_filtered, active_filters = apply_dashboard_filters(df, params)
    payload = build_dashboard_payload(df_filtered, total_unfiltered=len(df), active_filters=active_filters)
    app_state["dashboard_last_payload"] = payload
    return jsonify(payload)


@app.route("/api/case/<case_id>")
def get_case(case_id):
    df = app_state.get("df_scored")
    if df is None:
        return jsonify({"error": "Pipeline no ejecutado"}), 400

    mask = df["id_siniestro"].str.upper() == case_id.upper()
    if mask.sum() == 0:
        return jsonify({"error": f"Siniestro {case_id} no encontrado"}), 404

    row = df[mask].iloc[0]
    explanation = explain_single_case(row)
    return jsonify(explanation)


@app.route("/api/agent-status")
def agent_status():
    """Indica si ChatGPT (OpenAI) está configurado vía variables de entorno."""
    return jsonify({
        "openai_configured": is_openai_configured(),
        "openai_model": get_openai_model() if is_openai_configured() else None,
        "pipeline_ready": app_state.get("agent") is not None,
        "vercel": is_vercel_runtime(),
        "demo_mode": is_vercel_runtime(),
    })


@app.route("/api/agent-query", methods=["POST"])
def agent_query():
    agent = app_state.get("agent")
    if agent is None:
        return jsonify({"error": "Agente no inicializado. Ejecute el pipeline primero."}), 400

    data = request.get_json()
    question = data.get("question", "")
    if not question:
        return jsonify({"error": "Pregunta vacia"}), 400

    agent.set_extra_context(_build_agent_context())
    result = agent.query(question)
    if "datos" in result and result["datos"] is not None:
        if isinstance(result["datos"], list):
            for item in result["datos"]:
                for k, v in item.items():
                    if isinstance(v, float) and (v != v):
                        item[k] = None
        elif isinstance(result["datos"], dict):
            for k, v in result["datos"].items():
                if isinstance(v, float) and (v != v):
                    result["datos"][k] = None

    return jsonify(result)


# ── Export endpoints ─────────────────────────────────────────────────

@app.route("/api/export-powerbi", methods=["POST"])
def export_powerbi():
    try:
        df = app_state.get("df_scored")
        if df is None:
            return jsonify({"error": "Pipeline no ejecutado"}), 400

        output_path = os.path.join("data", "processed", "powerbi_export.xlsx")
        export_to_powerbi(app_state["datasets"], df, output_path)
        export_csv_for_powerbi(df)

        return jsonify({
            "status": "success",
            "message": "Exportacion completada",
            "excel_path": output_path,
            "csv_path": "data/processed/siniestros_scored.csv",
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download-powerbi")
def download_powerbi():
    path = os.path.join("data", "processed", "powerbi_export.xlsx")
    if os.path.exists(path):
        return send_file(path, as_attachment=True, download_name="powerbi_export.xlsx")
    return jsonify({"error": "Archivo no encontrado. Ejecute la exportacion primero."}), 404


@app.route("/api/download-csv")
def download_csv():
    path = os.path.join("data", "processed", "siniestros_scored.csv")
    if os.path.exists(path):
        return send_file(path, as_attachment=True, download_name="siniestros_scored.csv")
    return jsonify({"error": "Archivo no encontrado"}), 404


@app.route("/api/model-metrics")
def model_metrics():
    results = app_state.get("model_results")
    if results is None:
        snapshot = app_state.get("model_snapshot")
        if snapshot:
            return jsonify(snapshot)
        return jsonify({"error": "Modelo no entrenado"}), 400

    if is_vercel_runtime() or results.get("trained") or "model" not in results:
        return jsonify(results)

    metrics = get_model_metrics_summary(results)
    metrics["confusion_matrix"] = results.get("confusion_matrix")
    return jsonify(metrics)


@app.route("/api/nlp-summary")
def nlp_summary():
    results = app_state.get("nlp_results")
    if results is None:
        return jsonify({"error": "Analisis NLP no ejecutado"}), 400
    return jsonify(results)


@app.route("/api/status")
def get_status():
    db_info = (
        {"status": "ok", "type": "In-memory (Vercel)", "host": "demo"}
        if is_vercel_runtime()
        else test_connection()
    )
    return jsonify({
        "pipeline_status": app_state["pipeline_status"],
        "datasets_loaded": list(app_state["datasets"].keys()),
        "has_scored_data": app_state["df_scored"] is not None,
        "has_model": app_state["model_results"] is not None,
        "has_agent": app_state["agent"] is not None,
        "database": db_info,
        "vercel": is_vercel_runtime(),
        "openai_configured": is_openai_configured(),
    })


if __name__ == "__main__":
    init_database()
    db_info = test_connection()
    port = int(os.environ.get("FLASK_PORT", 5000))
    print(f"\n{'='*60}")
    print(f"  FraudIA Claims - Sistema de Deteccion de Fraude")
    print(f"  Servidor: http://localhost:{port}")
    print(f"  Base de datos: {db_info.get('type', '?')} ({db_info.get('host', 'local')})")
    print(f"  Status BD: {db_info.get('status', '?')}")
    print(f"{'='*60}\n")
    app.run(host="0.0.0.0", port=port, debug=True)
